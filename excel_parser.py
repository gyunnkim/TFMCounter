#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
from datetime import datetime
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다. 설치해주세요: pip install openpyxl")
    sys.exit(1)

def parse_game_data(worksheet, initial_date_str, year):
    """워크시트에서 게임 데이터를 파싱 (시트 중간에 날짜가 있는 경우도 처리)"""
    games = []
    current_date_str = initial_date_str
    
    # 플레이어 이름 찾기 (보통 3행에 있음)
    players = []
    for col in range(2, 5):  # B, C, D 열
        cell_value = worksheet.cell(row=3, column=col).value
        if cell_value and str(cell_value).strip():
            players.append(str(cell_value).strip())
    
    if len(players) < 2:
        return games
    
    # 게임 데이터 파싱
    current_row = 4
    max_row = worksheet.max_row
    
    while current_row <= max_row:
        cell_a = worksheet.cell(row=current_row, column=1).value
        
        # 날짜 셀인지 확인 (datetime 객체인 경우)
        if cell_a and hasattr(cell_a, 'strftime'):
            current_date_str = cell_a.strftime('%Y. %m. %d.')
            current_row += 1
            continue
        
        # 세트 정보 확인
        set_info = cell_a
        if not set_info or 'set' not in str(set_info).lower():
            current_row += 1
            continue
        
        # 기업 정보 (현재 행)
        corporations = []
        for col in range(2, len(players) + 2):
            corp = worksheet.cell(row=current_row, column=col).value
            corporations.append(str(corp).strip() if corp else "")
        
        # 맵과 점수 정보 (다음 행)
        current_row += 1
        if current_row > max_row:
            break
            
        map_name = ""
        scores = []
        megacredits = []
        
        # 맵 이름과 점수 파싱
        map_cell = worksheet.cell(row=current_row, column=1).value
        if map_cell:
            map_name = str(map_cell).strip()
        
        for col in range(2, len(players) + 2):
            score_cell = worksheet.cell(row=current_row, column=col).value
            if score_cell:
                score_str = str(score_cell)
                # 점수와 메가크레딧 분리 (예: "98(21)" -> 점수: 98, MC: 21)
                if '(' in score_str and ')' in score_str:
                    score = int(score_str.split('(')[0])
                    mc = int(score_str.split('(')[1].split(')')[0])
                else:
                    score = int(float(score_str)) if score_str.replace('.', '').isdigit() else 0
                    mc = 0
                scores.append(score)
                megacredits.append(mc)
            else:
                scores.append(0)
                megacredits.append(0)
        
        # 순위 계산 (점수 기준, 동점시 메가크레딧)
        player_results = []
        for i, player in enumerate(players):
            if i < len(scores):
                player_results.append({
                    'player': player,
                    'corporation': corporations[i] if i < len(corporations) else "",
                    'score': scores[i],
                    'megacredits': megacredits[i]
                })
        
        # 점수와 메가크레딧으로 정렬
        player_results.sort(key=lambda x: (x['score'], x['megacredits']), reverse=True)
        
        # 게임 데이터 생성
        game_results = []
        for rank, result in enumerate(player_results, 1):
            game_results.append({
                'playerId': hash(result['player']) % 1000000,  # 임시 ID
                'playerName': result['player'],
                'cubeColor': ['red', 'green', 'yellow', 'blue'][rank-1] if rank <= 4 else 'black',
                'corporation': result['corporation'],
                'score': result['score'],
                'megacredits': result['megacredits'],
                'rank': rank
            })
        
        # 맵 이름 정규화
        map_normalized = normalize_map_name(map_name)
        
        games.append({
            'date': current_date_str,
            'map': map_normalized,
            'results': game_results,
            'year': year
        })
        
        current_row += 1
    
    return games

def normalize_map_name(map_name):
    """맵 이름 정규화"""
    map_name = map_name.lower()
    if '타르시스' in map_name or 'tharsis' in map_name:
        return 'THARSIS'
    elif '헬라스' in map_name or 'hellas' in map_name:
        return 'HELLAS'
    elif '엘리시움' in map_name or 'elysium' in map_name:
        return 'ELYSIUM'
    elif '테라' in map_name or 'terra' in map_name or 'cimeria' in map_name:
        return 'TERRA CIMERIA'
    elif 'vastitas' in map_name or '바스티타스' in map_name:
        return 'VASTITAS BOREALIS'
    elif 'utopia' in map_name or '유토피아' in map_name:
        return 'UTOPIA PLANITIA'
    else:
        return 'THARSIS'  # 기본값

def parse_tfm_excel_files():
    """테라포밍 마스 Excel 파일들을 파싱해서 레거시 데이터 생성"""
    
    downloads_path = "/Users/kihokim/Downloads"
    excel_files = [
        ("20190222-24_TFM.xlsx", 2019),
        ("2020_TFM.xlsx", 2020), 
        ("2021_TFM.xlsx", 2021),
        ("2022_TFM.xlsx", 2022)
    ]
    
    all_players = {}
    all_games = []
    game_id_counter = 1
    
    print("🔍 테라포밍 마스 Excel 파일 파싱 시작...")
    
    for excel_file, year in excel_files:
        file_path = os.path.join(downloads_path, excel_file)
        if not os.path.exists(file_path):
            print(f"❌ 파일을 찾을 수 없습니다: {excel_file}")
            continue
            
        print(f"\n📊 파싱 중: {excel_file} ({year}년)")
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            
            # Online 탭 제외하고 처리
            for sheet_name in sheet_names:
                if 'online' in sheet_name.lower() or not sheet_name.strip():
                    continue
                    
                print(f"   📋 시트 처리: {sheet_name}")
                
                try:
                    worksheet = workbook[sheet_name]
                    
                    # 날짜 추출 (첫 번째 셀에서)
                    date_cell = worksheet.cell(row=1, column=1).value
                    if date_cell and hasattr(date_cell, 'strftime'):
                        date_str = date_cell.strftime('%Y. %m. %d.')
                    else:
                        date_str = f"{year}. 01. 01."
                    
                    # 게임 데이터 파싱
                    games = parse_game_data(worksheet, date_str, year)
                    
                    for game in games:
                        game['id'] = game_id_counter
                        all_games.append(game)
                        game_id_counter += 1
                        
                        # 플레이어 통계 업데이트
                        for result in game['results']:
                            player_name = result['playerName']
                            if player_name not in all_players:
                                all_players[player_name] = {
                                    'id': len(all_players) + 1,
                                    'name': player_name,
                                    'games': [],
                                    'stats': {
                                        'totalGames': 0,
                                        'totalScore': 0,
                                        'averageScore': 0,
                                        'wins': 0,
                                        'seconds': 0,
                                        'thirds': 0,
                                        'fourths': 0
                                    }
                                }
                            
                            player = all_players[player_name]
                            player['games'].append(result)
                            player['stats']['totalGames'] += 1
                            player['stats']['totalScore'] += result['score']
                            
                            if result['rank'] == 1:
                                player['stats']['wins'] += 1
                            elif result['rank'] == 2:
                                player['stats']['seconds'] += 1
                            elif result['rank'] == 3:
                                player['stats']['thirds'] += 1
                            elif result['rank'] == 4:
                                player['stats']['fourths'] += 1
                    
                    print(f"      ✅ {len(games)}개 게임 파싱 완료")
                        
                except Exception as e:
                    print(f"      ❌ 시트 파싱 오류: {e}")
                    continue
            
            workbook.close()
                    
        except Exception as e:
            print(f"❌ 파일 읽기 오류 ({excel_file}): {e}")
            continue
    
    # 평균 점수 계산
    for player in all_players.values():
        if player['stats']['totalGames'] > 0:
            player['stats']['averageScore'] = round(
                player['stats']['totalScore'] / player['stats']['totalGames'], 1
            )
    
    # 레거시 데이터 생성
    legacy_data = {
        'players': list(all_players.values()),
        'games': all_games,
        'exportDate': datetime.now().isoformat(),
        'version': '1.0',
        'description': '테라포밍 마스 레거시 데이터 (2019-2022)',
        'source': 'Excel 파일 파싱',
        'totalGames': len(all_games),
        'totalPlayers': len(all_players),
        'yearRange': '2019-2022'
    }
    
    # JSON 파일로 저장
    output_file = '/Users/kihokim/Documents/TFMCounter/terraforming_mars_legacy_2019-2022.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(legacy_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 레거시 데이터 생성 완료!")
    print(f"📁 파일 위치: {output_file}")
    print(f"👥 총 플레이어: {len(all_players)}명")
    print(f"🎮 총 게임: {len(all_games)}게임")
    
    # 플레이어별 통계 출력
    print(f"\n📊 플레이어별 통계:")
    sorted_players = sorted(all_players.values(), 
                          key=lambda x: (x['stats']['wins'], x['stats']['averageScore']), 
                          reverse=True)
    
    for i, player in enumerate(sorted_players, 1):
        stats = player['stats']
        print(f"   {i}. {player['name']}: {stats['totalGames']}게임, "
              f"{stats['wins']}승, 평균 {stats['averageScore']}점")
    
    return legacy_data

if __name__ == "__main__":
    parse_tfm_excel_files()
